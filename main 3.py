import os
import chardet
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def detect_encoding(file_path):
    with open(file_path, 'rb') as file:
        raw_data = file.read()
    return chardet.detect(raw_data)['encoding']


def process_csv_files(folder_path, frequency='Daily', rank_filter=1):
    try:
        # Step 1: Read all CSV files in the folder
        all_data = []
        for filename in os.listdir(folder_path):
            if filename.endswith('.csv'):
                file_path = os.path.join(folder_path, filename)
                try:
                    encoding = detect_encoding(file_path)
                    df = pd.read_csv(file_path, encoding=encoding)
                    df['Strategy'] = filename.replace('.csv', '')
                    all_data.append(df)
                except Exception as e:
                    print(f"Error reading file {filename}: {str(e)}")
                    continue

        if not all_data:
            print(f"No valid CSV files found in {folder_path}")
            return pd.DataFrame(), pd.DataFrame()

        # Combine all dataframes
        combined_df = pd.concat(all_data, ignore_index=True)

        # Step 2: Group by entry date and sum profit
        combined_df.columns = combined_df.columns.str.lower()

        # Convert entry_date to datetime, handling both "YYYY-MM-DD" and "DD/MM/YY" formats
        combined_df['entry_date'] = pd.to_datetime(combined_df['entry_date'], format='%Y-%m-%d', errors='coerce')
        mask = combined_df['entry_date'].isna()
        combined_df.loc[mask, 'entry_date'] = pd.to_datetime(combined_df.loc[mask, 'entry_date'], format='%d/%m/%y',
                                                             errors='coerce')

        # Drop rows with invalid dates
        combined_df = combined_df.dropna(subset=['entry_date'])

        # Apply frequency
        if frequency.capitalize() == 'Weekly':
            combined_df['period'] = combined_df['entry_date'] - pd.to_timedelta(combined_df['entry_date'].dt.dayofweek,
                                                                                unit='D')
        elif frequency.capitalize() == 'Monthly':
            combined_df['period'] = combined_df['entry_date'].dt.to_period('M').dt.to_timestamp()
        elif frequency.capitalize() == 'Quarterly':
            combined_df['period'] = combined_df['entry_date'].dt.to_period('Q').dt.to_timestamp()
        else:  # Daily
            combined_df['period'] = combined_df['entry_date']

        # Group by period and strategy, then sum profits
        grouped_df = combined_df.groupby(['period', 'strategy'])['profit'].sum().reset_index()

        # Pivot the dataframe
        pivot_df = grouped_df.pivot(index='period', columns='strategy', values='profit').reset_index()
        pivot_df = pivot_df.fillna(0)

        # Sort by period
        pivot_df = pivot_df.sort_values('period')

        # Step 4: Select strategy based on previous period's performance
        strategy_columns = pivot_df.columns.drop(['period'])
        selected_strategies = []
        selected_profits = []

        for i in range(len(pivot_df)):
            if i == 0:
                # For the first period, use its own data
                profits = pivot_df.iloc[0][strategy_columns]
                # Ensure all values are numeric
                profits = profits.apply(pd.to_numeric, errors='coerce')
                selected_strategy = profits.idxmax()
            else:
                # For subsequent periods, use previous period's data
                profits = pivot_df.iloc[i - 1][strategy_columns]
                # Ensure all values are numeric
                profits = profits.apply(pd.to_numeric, errors='coerce')
                sorted_strategies = profits.sort_values(ascending=False)
                selected_strategy = sorted_strategies.index[
                    min(rank_filter - 1, len(sorted_strategies) - 1)]  # Ensure we don't go out of bounds

            selected_strategies.append(selected_strategy)
            selected_profits.append(pivot_df.iloc[i][selected_strategy])

        # Create final dataframe
        final_df = pd.DataFrame({
            'Date': pivot_df['period'],
            'Strategy': selected_strategies,
            'Profit': selected_profits
        })

        # Ensure Date is in the correct format
        final_df['Date'] = pd.to_datetime(final_df['Date']).dt.strftime('%Y-%m-%d')

        return final_df, pivot_df

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame(), pd.DataFrame()


def save_to_excel(df, file_name):
    # Create 'output' folder if it doesn't exist
    if not os.path.exists('output'):
        os.makedirs('output')

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Write the DataFrame to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Apply formatting to header row
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # Apply general formatting to all cells
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    full_path = os.path.join('output', file_name)
    wb.save(full_path)
    print(f"Data saved to {full_path}")


def main():
    # folder_path = 'path/to/your/csv/files'
    folder_path = 'data/STRATEGIES/'
    frequencies = ['Daily', 'Weekly', 'Monthly', 'Quarterly']
    rank_filter = 1  # Change this to the desired rank

    for frequency in frequencies:
        final_df, full_df = process_csv_files(folder_path, frequency=frequency, rank_filter=rank_filter)

        # Save the main report
        save_to_excel(final_df, f'best_strategy_{frequency.lower()}.xlsx')

        # Save the full report
        save_to_excel(full_df, f'full_analysis_{frequency.lower()}.xlsx')


if __name__ == "__main__":
    main()
